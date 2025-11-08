"""Service layer to expose password-manager functionality using SQLAlchemy models

This adapts the original passwd_manager/database.py logic to use the main
app SQLAlchemy instance so the tables live inside the BilancioFamiliare DB.
"""
from app import db
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives import hashes
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
import base64
from app.models.PasswdCredential import PasswdCredential
from app.models.PasswdSecurityConfig import PasswdSecurityConfig
from typing import Optional
import os
import tempfile
import openpyxl
import logging

logger = logging.getLogger(__name__)

# Module-level cipher (similar behaviour to original project)
_cipher: Optional[Fernet] = None


def derive_key_from_password(password: str, salt: bytes) -> bytes:
    kdf = PBKDF2HMAC(algorithm=hashes.SHA256(), length=32, salt=salt, iterations=100000)
    key = base64.urlsafe_b64encode(kdf.derive(password.encode()))
    return key


def initialize_encryption(password: str) -> bool:
    """Initialize cipher using password and the stored security config.

    Returns True when the provided password unlocks the stored `test_encrypted`.
    """
    global _cipher
    try:
        cfg = PasswdSecurityConfig.query.filter_by(id=1).first()
        if not cfg or not cfg.salt or not cfg.test_encrypted:
            return False

        key = derive_key_from_password(password, cfg.salt)
        test_cipher = Fernet(key)
        try:
            test_data = test_cipher.decrypt(cfg.test_encrypted.encode()).decode()
            if test_data == 'test_string':
                _cipher = test_cipher
                return True
        except Exception:
            return False
    except Exception as e:
        logger.exception('initialize_encryption error: %s', e)
        return False


def is_initialized() -> bool:
    return _cipher is not None


def has_security_config() -> bool:
    cfg = PasswdSecurityConfig.query.filter_by(id=1).first()
    return cfg is not None and cfg.test_encrypted is not None


def encrypt_data(data: str) -> str:
    if not data or not _cipher:
        return ''
    try:
        return _cipher.encrypt(data.encode()).decode()
    except Exception:
        return ''


def decrypt_data(encrypted_data: str) -> str:
    if not encrypted_data or not _cipher:
        return ''
    try:
        return _cipher.decrypt(encrypted_data.encode()).decode()
    except Exception:
        return ''


def init_database_if_needed():
    """Create tables (if needed). This relies on db.create_all() being called
    from the app factory; kept for compatibility with tests or direct calls."""
    db.create_all()


def get_all_credentials():
    if not is_initialized():
        return []
    rows = PasswdCredential.query.order_by(PasswdCredential.categoria, PasswdCredential.servizio).all()
    return [r.to_decrypted_dict(decrypt_data) for r in rows]


def add_credential(categoria, servizio, utenza, password, altro):
    if not is_initialized():
        return None
    pw_enc = encrypt_data(password or '')
    altro_enc = encrypt_data(altro or '')
    c = PasswdCredential(categoria=categoria, servizio=servizio, utenza=utenza or '', password=pw_enc, altro=altro_enc)
    db.session.add(c)
    db.session.commit()
    return c.id


def delete_credential(credential_id):
    c = PasswdCredential.query.get(credential_id)
    if not c:
        return False
    db.session.delete(c)
    db.session.commit()
    return True


def update_credential(credential_id, categoria, servizio, utenza, password, altro):
    if not is_initialized():
        return False
    c = PasswdCredential.query.get(credential_id)
    if not c:
        return False
    c.categoria = categoria
    c.servizio = servizio
    c.utenza = utenza or ''
    c.password = encrypt_data(password or '')
    c.altro = encrypt_data(altro or '')
    db.session.commit()
    return True


def search_credentials(query, category_filter=None):
    if not is_initialized():
        return []
    q = PasswdCredential.query
    if query:
        like = f"%{query.lower()}%"
        q = q.filter(db.or_(db.func.lower(PasswdCredential.categoria).like(like), db.func.lower(PasswdCredential.servizio).like(like)))
    if category_filter:
        q = q.filter_by(categoria=category_filter)
    rows = q.order_by(PasswdCredential.categoria, PasswdCredential.servizio).all()
    return [r.to_decrypted_dict(decrypt_data) for r in rows]


def get_categories():
    rows = db.session.query(PasswdCredential.categoria).distinct().order_by(PasswdCredential.categoria).all()
    return [r[0] for r in rows if r[0]]


def get_credential_by_id(credential_id):
    c = PasswdCredential.query.get(credential_id)
    if not c:
        return None
    return {
        'id': c.id,
        'CATEGORIA': c.categoria,
        'SERVIZIO': c.servizio,
        'UTENZA': c.utenza,
        'PASSWORD': c.password,
        'ALTRO': c.altro
    }


def get_credential_by_id_decrypted(credential_id):
    c = PasswdCredential.query.get(credential_id)
    if not c:
        return None
    return c.to_decrypted_dict(decrypt_data)


def export_to_xlsx(protection_password=None):
    """Export all decrypted credentials to a temporary xlsx file.

    If msoffcrypto is installed and protection_password is provided, it will try
    to protect the file.
    """
    try:
        data = get_all_credentials()
        rows = [(r['CATEGORIA'], r['SERVIZIO'], r['UTENZA'], r['PASSWORD'], r['ALTRO']) for r in data]

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Password Database'
        headers = ['CATEGORIA', 'SERVIZIO', 'UTENZA', 'PASSWORD', 'ALTRO']
        for i, h in enumerate(headers, 1):
            ws.cell(row=1, column=i, value=h)
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(tmp.name)

        # try to protect with msoffcrypto if requested
        if protection_password:
            try:
                import msoffcrypto
                import io
                with open(tmp.name, 'rb') as f:
                    office_file = msoffcrypto.OfficeFile(f)
                    # produce protected file
                    protected_tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
                    with open(protected_tmp.name, 'wb') as out_f:
                        office_file.encrypt(protection_password, out_f)
                    os.unlink(tmp.name)
                    return protected_tmp.name
            except Exception:
                logger.exception('msoffcrypto not available or encryption failed')
                return tmp.name

        return tmp.name
    except Exception:
        logger.exception('export_to_xlsx failed')
        return None
